from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QPushButton, QTableWidget,
                             QTableWidgetItem, QDialog, QFormLayout, QLineEdit,
                             QDateTimeEdit, QMessageBox, QHBoxLayout, QHeaderView,
                             QDialogButtonBox, QComboBox, QSpinBox)
from PyQt6.QtCore import Qt, QDateTime
from database import SessionLocal
from models import Passenger, Train, Ticket, Station
import openpyxl
from datetime import datetime
from .ticket_dialog import TicketDialog


class MainWindow(QWidget):
    def __init__(self, parent=None):
        super().__init__()
        self.parent = parent
        self.setup_ui()
        self.load_data()

    def setup_ui(self):
        layout = QVBoxLayout()
        self.setLayout(layout)

        # Создаем таблицу
        self.table = QTableWidget()
        headers = ["ID", "ФИО", "Паспорт", "Название поезда", "Место", "Станция отправления", "Станция прибытия",
                   "Время отправления", "Время прибытия", "ФИО Кассира"]
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)
        self.table.verticalHeader().setVisible(False)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        # Настраиваем внешний вид таблицы
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: white;
                alternate-background-color: #F5F5F5;
                border: 1px solid #E0E0E0;
                color: black;
                gridline-color: #E0E0E0;
            }
            QTableWidget::item {
                padding: 5px;
                color: black;
            }
            QHeaderView::section {
                background-color: #1976D2;
                color: white;
                padding: 8px;
                border: none;
                font-weight: bold;
            }
            QTableWidget::item:selected {
                background-color: #BBDEFB;
                color: black;
            }
        """)

        # Устанавливаем растягивание колонок
        header = self.table.horizontalHeader()
        for i in range(len(headers)):
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)

        # Кнопки управления
        button_layout = QHBoxLayout()

        buttons_data = [
            ("Добавить", "#4CAF50", self.add_passenger),
            ("Экспорт в Excel", "#FF9800", self.export_to_excel),
            ("Обновить", "#9C27B0", self.load_data)
        ]

        if self.parent.user:
            with SessionLocal() as db:
                if self.parent.user.is_admin:
                    buttons_data.append(("Редактировать", "#2196F3", self.edit_passenger))
                    buttons_data.append(("Удалить", "#F44336", self.delete_passenger))

        for text, color, callback in buttons_data:
            button = QPushButton(text)
            button.clicked.connect(callback)
            button.setStyleSheet(f"""
                QPushButton {{
                    background-color: {color};
                    color: white;
                    border: none;
                    padding: 10px;
                    border-radius: 4px;
                    min-width: 120px;
                    margin: 10px;
                }}
                QPushButton:hover {{
                    background-color: {color}DD;
                }}
                QPushButton:pressed {{
                    background-color: {color}AA;
                }}
            """)
            button_layout.addWidget(button)

        layout.addLayout(button_layout)
        layout.addWidget(self.table)

    def load_data(self):
        with SessionLocal() as db:
            tickets = db.query(Ticket).all()
            self.table.setRowCount(len(tickets))

            for i, ticket in enumerate(tickets):
                full_name = ' '.join(
                    [ticket.passenger.first_name, ticket.passenger.middle_name, ticket.passenger.last_name])
                passport = ' '.join(map(str, [ticket.passenger.series_passport, ticket.passenger.number_passport]))
                cashier_full_name = ' '.join([ticket.cashier.lastname, ticket.cashier.firstname,
                                              ticket.cashier.middle_name]) if ticket.cashier else ""

                items = [
                    QTableWidgetItem(str(ticket.id)),
                    QTableWidgetItem(full_name),
                    QTableWidgetItem(passport),
                    QTableWidgetItem(ticket.train.train_name),
                    QTableWidgetItem(str(ticket.seat_number)),
                    QTableWidgetItem(ticket.departure_station.name_station),
                    QTableWidgetItem(ticket.arrival_station.name_station),
                    QTableWidgetItem(ticket.departure_time.strftime("%Y-%m-%d %H:%M")),
                    QTableWidgetItem(ticket.arrival_time.strftime("%Y-%m-%d %H:%M")),
                    QTableWidgetItem(cashier_full_name),
                ]

                for j, item in enumerate(items):
                    self.table.setItem(i, j, item)
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)

    def add_passenger(self):
        dialog = TicketDialog(self, self.parent.user)
        dialog.exec()
        self.load_data()

    def edit_passenger(self):
        current_row = self.table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Предупреждение", "Выберите пассажира для редактирования")
            return

        with SessionLocal() as db:
            ticket_id = int(self.table.item(current_row, 0).text())
            ticket = db.query(Ticket).filter_by(id=ticket_id).first()

            if not ticket:
                QMessageBox.warning(self, "Предупреждение", "Пассажир не найден")
                return

            dialog = TicketDialog(self, self.parent.user, ticket)

        dialog.exec()
        self.load_data()

    def delete_passenger(self):
        current_row = self.table.currentRow()
        if current_row < 0:
            QMessageBox.warning(self, "Предупреждение", "Выберите пассажира для удаления")
            return

        with SessionLocal() as db:
            ticket_id = int(self.table.item(current_row, 0).text())
            ticket = db.query(Ticket).filter_by(id=ticket_id).first()

            if not ticket:
                QMessageBox.warning(self, "Предупреждение", "Пассажир не найден")
                return

            reply = QMessageBox.question(self, "Подтверждение",
                                         "Вы уверены, что хотите удалить этого пассажира?",
                                         QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

            if reply == QMessageBox.StandardButton.Yes:
                db.delete(ticket)
                db.commit()
                self.load_data()

    def export_to_excel(self):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Пассажиры"

            # Записываем заголовки
            headers = ["ID", "ФИО", "Паспорт", "Название поезда", "Место", "Станция отправления", "Станция прибытия",
                       "Время отправления", "Время прибытия"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Записываем данные
            with SessionLocal() as db:
                tickets = db.query(Ticket).all()
                for row, ticket in enumerate(tickets, 2):
                    full_name = ' '.join(
                        [ticket.passenger.first_name, ticket.passenger.middle_name, ticket.passenger.last_name])
                    passport = ' '.join(map(str, [ticket.passenger.series_passport, ticket.passenger.number_passport]))

                    values = [
                        ticket.id,
                        full_name,
                        passport,
                        ticket.train.train_name,
                        ticket.seat_number,
                        ticket.departure_station.name_station,
                        ticket.arrival_station.name_station,
                        ticket.departure_time.strftime("%Y-%m-%d %H:%M"),
                        ticket.arrival_time.strftime("%Y-%m-%d %H:%M"),
                    ]
                    for col, value in enumerate(values, 1):
                        ws.cell(row=row, column=col, value=str(value))

            # Сохраняем файл
            filename = f"passengers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)
            QMessageBox.information(self, "Успех", f"Данные экспортированы в файл {filename}")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при экспорте данных: {str(e)}")
